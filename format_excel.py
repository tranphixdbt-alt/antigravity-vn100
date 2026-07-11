import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

df = pd.read_csv('/Users/macos/Desktop/VN100_Valuation_Results.csv')

wb = Workbook()
ws = wb.active
ws.title = "VN100 Valuation"

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin', color='E0E0E0'),
                     right=Side(style='thin', color='E0E0E0'),
                     top=Side(style='thin', color='E0E0E0'),
                     bottom=Side(style='thin', color='E0E0E0'))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

number_format = '#,##0'
percent_format = '0.00%'

headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = thin_border
        
    if headers.get('Current Price'): row[headers['Current Price']-1].number_format = number_format
    if headers.get('Intrinsic FV'): row[headers['Intrinsic FV']-1].number_format = number_format
    if headers.get('Relative FV'): row[headers['Relative FV']-1].number_format = number_format
    if headers.get('Blended FV'): row[headers['Blended FV']-1].number_format = number_format
    
    upside_idx = headers.get('Upside (%)')
    if upside_idx and row[upside_idx-1].value is not None:
        try:
            row[upside_idx-1].value = float(row[upside_idx-1].value) / 100.0
            row[upside_idx-1].number_format = percent_format
        except: pass
        
    mos_idx = headers.get('MOS Target (%)')
    if mos_idx and row[mos_idx-1].value is not None:
        try:
            row[mos_idx-1].value = float(row[mos_idx-1].value) / 100.0
            row[mos_idx-1].number_format = percent_format
        except: pass
        
    rec_idx = headers.get('Recommendation')
    if rec_idx and row[rec_idx-1].value:
        rec_val = str(row[rec_idx-1].value).upper()
        if 'BUY' in rec_val:
            row[rec_idx-1].font = Font(color="00B050", bold=True)
        elif 'SELL' in rec_val:
            row[rec_idx-1].font = Font(color="FF0000", bold=True)
        elif 'TRIM' in rec_val or 'HOLD' in rec_val:
            row[rec_idx-1].font = Font(color="E26B0A", bold=True)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except: pass
    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

# Filter
ws.auto_filter.ref = ws.dimensions

# Freeze top row
ws.freeze_panes = "A2"

output_path = '/Users/macos/Desktop/VN100_Valuation_Pro.xlsx'
wb.save(output_path)
print("Saved to", output_path)
