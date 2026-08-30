import logging
import datetime
from collections import Counter
from pathlib import Path
import pandas as pd
import gspread
from gspread_dataframe import set_with_dataframe
from sqlalchemy import func
from sqlalchemy.orm import Session
from valuation.db.models import (
    DailySignal,
    FinancialsQuarterly,
    MacroSeries,
    PricesDaily,
    Ticker,
)
from valuation.db.session import SessionLocalRead
from valuation.config import settings
from valuation.engine.ttm_helper import (
    build_vcb_current_financials,
    build_fpt_current_financials,
    build_hpg_current_financials,
    build_ssi_current_financials,
    build_dgc_current_financials
)
from valuation.engine.consensus_helper import get_consensus_stats
from valuation.analysis.ai_insight import generate_ai_insight

logger = logging.getLogger(__name__)

def export_daily_signals_to_gsheets(trade_date: datetime.date = None, db: Session = None):
    """
    Xuất bảng DailySignal trong ngày ra Google Sheets với giao diện chuyên nghiệp, các chỉ số bổ sung và nhận định AI.
    """
    if not settings.google_service_account_json or not settings.google_sheet_master_id:
        logger.warning("Google Sheets config missing. Skipping export.")
        return {"status": "skipped", "reason": "Missing GS config"}
        
    close_db = False
    if db is None:
        db = SessionLocalRead()
        close_db = True
        
    if trade_date is None:
        trade_date = datetime.date.today()
        
    try:
        # Fetch data
        signals = db.query(DailySignal).filter(DailySignal.trade_date == trade_date).all()
        if not signals:
            logger.info("No daily signals found for export.")
            return {"status": "success", "exported_rows": 0}
            
        # Tương thích với các hàm build_current_financials
        fin_builders = {
            "VCB": build_vcb_current_financials,
            "FPT": build_fpt_current_financials,
            "HPG": build_hpg_current_financials,
            "SSI": build_ssi_current_financials,
            "DGC": build_dgc_current_financials
        }
        
        # Map sector sang tiếng Việt chuyên nghiệp
        SECTOR_MAP = {
            "Banks": "Ngân hàng",
            "Technology": "Công nghệ",
            "Chemicals": "Hóa chất",
            "Securities": "Chứng khoán",
            "Thép / vật liệu": "Thép & Vật liệu"
        }
        
        data = []
        for s in signals:
            ticker_obj = db.query(Ticker).filter(Ticker.ticker == s.ticker).first()
            sector = ticker_obj.sector if ticker_obj else "Unknown"
            display_sector = SECTOR_MAP.get(sector, sector)
            
            close_price = float(s.close_price) if s.close_price else None
            fv_fast = float(s.fair_value_fast) if s.fair_value_fast else None
            
            # Tính toán các chỉ số cơ bản
            pe = None
            pb = None
            roe = None
            
            if s.ticker in fin_builders and close_price:
                try:
                    fin_data = fin_builders[s.ticker](db)
                    equity = float(fin_data.get("total_equity") or 0.0)
                    net_inc = float(fin_data.get("net_income") or 0.0)
                    shares = float(fin_data.get("shares_outstanding") or 0.0)
                    
                    if equity > 0:
                        roe = net_inc / equity
                    if shares > 0:
                        eps = net_inc / shares
                        bvps = equity / shares
                        if eps > 0:
                            pe = close_price / eps
                        if bvps > 0:
                            pb = close_price / bvps
                except Exception as e:
                    logger.warning(f"Error computing metrics for {s.ticker}: {e}")
            
            # Lấy thông tin đồng thuận CTCK (Consensus)
            consensus_target = None
            consensus_dev = None
            try:
                c_stats = get_consensus_stats(s.ticker, s.trade_date, db)
                if c_stats["median"] is not None:
                    consensus_target = float(c_stats["median"])
                    if fv_fast:
                        consensus_dev = (fv_fast - consensus_target) / consensus_target
            except Exception as e:
                logger.warning(f"Error fetching consensus for {s.ticker}: {e}")
                
            # Xác định khuyến nghị (Rating)
            rating = "THEO DÕI"
            if s.upside is not None and s.margin_of_safety is not None:
                upside_val = float(s.upside)
                mos_val = float(s.margin_of_safety)
                if upside_val > mos_val:
                    rating = "MUA"
                elif upside_val < 0:
                    rating = "BÁN"
                else:
                    rating = "THEO DÕI"
                    
            # Gọi DeepSeek tạo nhận định AI (AI Insight)
            ai_insight = generate_ai_insight(
                ticker=s.ticker,
                sector=display_sector,
                close_price=close_price,
                fair_value=fv_fast,
                upside=s.upside,
                flags=s.flags,
                roe=roe,
                pe=pe,
                pb=pb,
                consensus_target=consensus_target
            )
            
            data.append({
                "Mã CK": s.ticker,
                "Ngành": display_sector,
                "Ngày GD": s.trade_date.isoformat(),
                "Thị giá": close_price,
                "FV Nhịp Nhanh": fv_fast,
                "Upside": float(s.upside) if s.upside is not None else None,
                "Biên An Toàn": float(s.margin_of_safety) if s.margin_of_safety is not None else None,
                "Khuyến nghị": rating,
                "Điểm Conviction": float(s.conviction_score) if s.conviction_score is not None else None,
                "P/E": pe,
                "P/B": pb,
                "ROE": roe,
                "Định giá Consensus": consensus_target,
                "Độ lệch Consensus": consensus_dev,
                "Cờ Cảnh Báo": ", ".join(s.flags) if s.flags else "OK",
                "Nhận định AI": ai_insight
            })
            
        df = pd.DataFrame(data)
        
        # Sắp xếp theo Điểm Conviction giảm dần
        df = df.sort_values(by="Điểm Conviction", ascending=False)
        
        # Authenticate and Push
        gc = gspread.service_account(filename=settings.google_service_account_json)
        sh = gc.open_by_key(settings.google_sheet_master_id)
        
        try:
            worksheet = sh.worksheet("Daily_Screener")
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sh.add_worksheet(title="Daily_Screener", rows="100", cols="20")
            
        # Clear old data and insert new
        worksheet.clear()
        set_with_dataframe(worksheet, df)
        
        # === Định dạng Google Sheets chuyên nghiệp ===
        num_rows = len(df) + 1 # bao gồm cả header
        
        # 1. Định dạng header (Dòng 1 từ cột A đến P)
        worksheet.format('A1:P1', {
            'backgroundColor': {
                'red': 0.12,
                'green': 0.23,
                'blue': 0.35
            },
            'horizontalAlignment': 'CENTER',
            'verticalAlignment': 'MIDDLE',
            'textFormat': {
                'foregroundColor': {
                    'red': 1.0,
                    'green': 1.0,
                    'blue': 1.0
                },
                'fontSize': 10,
                'bold': True
            }
        })
        
        # 2. Định dạng font chữ chung cho data (A đến P)
        worksheet.format(f'A2:P{num_rows}', {
            'textFormat': {
                'fontSize': 10
            },
            'verticalAlignment': 'MIDDLE'
        })
        
        # 3. Định dạng căn lề và định dạng số cho từng cột
        worksheet.format(f'A2:A{num_rows}', {'horizontalAlignment': 'CENTER'})
        worksheet.format(f'B2:B{num_rows}', {'horizontalAlignment': 'LEFT'})
        worksheet.format(f'C2:C{num_rows}', {'horizontalAlignment': 'CENTER'})
        
        # Thị giá & FV Nhịp Nhanh: Số nguyên có phân cách hàng nghìn
        worksheet.format(f'D2:E{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '#,##0'}
        })
        
        # Upside & Biên An Toàn: Phần trăm 1 chữ số thập phân
        worksheet.format(f'F2:G{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}
        })
        
        # Khuyến nghị: Căn giữa, chữ in đậm
        worksheet.format(f'H2:H{num_rows}', {
            'horizontalAlignment': 'CENTER',
            'textFormat': {'bold': True}
        })
        
        # Điểm Conviction: Số thực 1 chữ số thập phân
        worksheet.format(f'I2:I{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '0.0'}
        })
        
        # P/E & P/B: Số thực 2 chữ số thập phân
        worksheet.format(f'J2:K{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '0.00'}
        })
        
        # ROE: Phần trăm 1 chữ số thập phân
        worksheet.format(f'L2:L{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}
        })
        
        # Định giá Consensus: Số nguyên có phân cách hàng nghìn
        worksheet.format(f'M2:M{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '#,##0'}
        })
        
        # Độ lệch Consensus: Phần trăm 1 chữ số thập phân
        worksheet.format(f'N2:N{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}
        })
        
        worksheet.format(f'O2:O{num_rows}', {'horizontalAlignment': 'LEFT'})
        worksheet.format(f'P2:P{num_rows}', {'horizontalAlignment': 'LEFT'}) # Cột P: Nhận định AI
        
        # 4. Tô màu có điều kiện cho cột Khuyến nghị
        for idx, row in df.iterrows():
            row_num = idx + 2
            rating_val = row["Khuyến nghị"]
            cell_range = f"H{row_num}"
            if rating_val == "MUA":
                worksheet.format(cell_range, {
                    'textFormat': {
                        'bold': True,
                        'foregroundColor': {'red': 0.1, 'green': 0.5, 'blue': 0.1} # Xanh lá đậm
                    }
                })
            elif rating_val == "BÁN":
                worksheet.format(cell_range, {
                    'textFormat': {
                        'bold': True,
                        'foregroundColor': {'red': 0.8, 'green': 0.1, 'blue': 0.1} # Đỏ đậm
                    }
                })
            else: # THEO DÕI
                worksheet.format(cell_range, {
                    'textFormat': {
                        'bold': True,
                        'foregroundColor': {'red': 0.5, 'green': 0.5, 'blue': 0.5} # Xám
                    }
                })
                
        logger.info(f"Successfully exported {len(df)} rows to Google Sheets.")
        return {"status": "success", "exported_rows": len(df)}
        
    except Exception as e:
        logger.error(f"Failed to export to Google Sheets: {e}")
        return {"status": "error", "error": str(e)}
    finally:
        if close_db:
            db.close()


def update_single_ticker_to_gsheets(
    ticker: str, 
    curr_price: float, 
    blended_fv: float, 
    greeks: dict, 
    qc_flags: list, 
    db: Session,
    ai_insight: str = None,
    allow_ai_call: bool = True,
):
    """
    Cập nhật hoặc chèn mới kết quả định giá của một mã CK đơn lẻ vào Google Sheets 'Daily_Screener'.
    Đảm bảo chống trùng lặp: nếu đã có thì ghi đè dòng cũ, nếu chưa có thì append dòng mới.
    """
    if not settings.google_service_account_json or not settings.google_sheet_master_id:
        logger.warning("Google Sheets config missing. Skipping single ticker update.")
        return {"status": "skipped", "reason": "Missing GS config"}
        
    try:
        # 1. Xác định Ngành và map tiếng Việt
        ticker_obj = db.query(Ticker).filter(Ticker.ticker == ticker).first()
        sector = ticker_obj.sector if ticker_obj else "Unknown"
        
        SECTOR_MAP = {
            "Banks": "Ngân hàng",
            "Technology": "Công nghệ",
            "Chemicals": "Hóa chất",
            "Securities": "Chứng khoán",
            "Thép / vật liệu": "Thép & Vật liệu"
        }
        display_sector = SECTOR_MAP.get(sector, sector)
        if not display_sector or display_sector == "":
            BANK_TICKERS = [
                "VCB", "ACB", "CTG", "BID", "TCB", "MBB", "VPB", "STB", "HDB", "VIB", "TPB", "MSB", "SHB", "OCB",
                "ABB", "BAB", "BVB", "EIB", "KLB", "LPB", "NAB", "NVB", "PGB", "SGB", "SSB", "VAB", "VBB"
            ]
            if ticker in BANK_TICKERS:
                display_sector = "Ngân hàng"
            else:
                display_sector = "Unknown"

        # 2. Tính toán ROE, P/E, P/B
        pe = None
        pb = None
        roe = None
        
        # Load builders
        fin_builders = {
            "VCB": build_vcb_current_financials,
            "FPT": build_fpt_current_financials,
            "HPG": build_hpg_current_financials,
            "SSI": build_ssi_current_financials,
            "DGC": build_dgc_current_financials
        }
        
        BANK_TICKERS = [
            "VCB", "ACB", "CTG", "BID", "TCB", "MBB", "VPB", "STB", "HDB", "VIB", "TPB", "MSB", "SHB", "OCB",
            "ABB", "BAB", "BVB", "EIB", "KLB", "LPB", "NAB", "NVB", "PGB", "SGB", "SSB", "VAB", "VBB"
        ]
        
        builder_fn = None
        if ticker in BANK_TICKERS:
            builder_fn = build_vcb_current_financials
        elif ticker in fin_builders:
            builder_fn = fin_builders[ticker]
            
        if builder_fn and curr_price:
            try:
                fin_data = builder_fn(db, ticker) if ticker in BANK_TICKERS else builder_fn(db)
                equity = float(fin_data.get("total_equity") or 0.0)
                net_inc = float(fin_data.get("net_income") or 0.0)
                shares = float(fin_data.get("shares_outstanding") or 0.0)
                
                if equity > 0:
                    roe = net_inc / equity
                if shares > 0:
                    eps = net_inc / shares
                    bvps = equity / shares
                    if eps > 0:
                        pe = curr_price / eps
                    if bvps > 0:
                        pb = curr_price / bvps
            except Exception as e:
                logger.warning(f"Error computing metrics for single {ticker}: {e}")
                
        # 3. Lấy Consensus target
        consensus_target = None
        consensus_dev = None
        trade_date = datetime.date.today()
        try:
            c_stats = get_consensus_stats(ticker, trade_date, db)
            if c_stats["median"] is not None:
                consensus_target = float(c_stats["median"])
                if blended_fv:
                    consensus_dev = (blended_fv - consensus_target) / consensus_target
        except Exception as e:
            logger.warning(f"Error fetching consensus for single {ticker}: {e}")
            
        # 4. Tính toán Upside & MOS (Biên an toàn) & Khuyến nghị
        upside = None
        mos = None
        if curr_price > 0:
            upside = (blended_fv - curr_price) / curr_price
            mos = 1.0 - (curr_price / blended_fv) if blended_fv > 0 else 0.0
            
        rating = "THEO DÕI"
        if upside is not None and mos is not None:
            if upside > mos:
                rating = "MUA"
            elif upside < 0:
                rating = "BÁN"
            else:
                rating = "THEO DÕI"
                
        # 5. UI ưu tiên dùng lại báo cáo đã kiểm chứng để không tạo API call thứ hai.
        if not ai_insight and allow_ai_call:
            ai_insight = generate_ai_insight(
                ticker=ticker,
                sector=display_sector,
                close_price=curr_price,
                fair_value=blended_fv,
                upside=upside,
                flags=qc_flags,
                roe=roe,
                pe=pe,
                pb=pb,
                consensus_target=consensus_target,
            )
        if not ai_insight:
            ai_insight = "Chưa sinh báo cáo AI đã kiểm chứng cho mã này."
        
        # 6. Chuẩn bị hàng dữ liệu để ghi vào Sheet
        row_data = [
            ticker,
            display_sector,
            trade_date.isoformat(),
            curr_price,
            blended_fv,
            upside,
            mos,
            rating,
            100.0, # Điểm Conviction (mặc định)
            pe,
            pb,
            roe,
            consensus_target,
            consensus_dev,
            ", ".join(qc_flags) if qc_flags else "OK",
            ai_insight
        ]
        
        # 7. Kết nối Google Sheets
        gc = gspread.service_account(filename=settings.google_service_account_json)
        sh = gc.open_by_key(settings.google_sheet_master_id)
        
        try:
            worksheet = sh.worksheet("Daily_Screener")
        except gspread.exceptions.WorksheetNotFound:
            worksheet = sh.add_worksheet(title="Daily_Screener", rows="100", cols="20")
            headers = [
                "Mã CK", "Ngành", "Ngày GD", "Thị giá", "FV Nhịp Nhanh", "Upside", "Biên An Toàn", 
                "Khuyến nghị", "Điểm Conviction", "P/E", "P/B", "ROE", "Định giá Consensus", 
                "Độ lệch Consensus", "Cờ Cảnh Báo", "Nhận định AI"
            ]
            worksheet.append_row(headers)
            
        # Đọc toàn bộ ticker hiện có ở cột A
        tickers_col = worksheet.col_values(1)
        
        # Tìm xem ticker đã tồn tại chưa
        row_index = -1
        for idx, val in enumerate(tickers_col):
            if val.strip().upper() == ticker.upper():
                row_index = idx + 1
                break
                
        if row_index > -1:
            cell_range = f"A{row_index}:P{row_index}"
            worksheet.update(cell_range, [row_data])
            logger.info(f"Updated ticker {ticker} at row {row_index} in Google Sheets.")
        else:
            worksheet.append_row(row_data)
            row_index = len(tickers_col) + 1
            logger.info(f"Appended ticker {ticker} as a new row in Google Sheets.")
            
        # 8. Định dạng lại dòng vừa được ghi/cập nhật
        worksheet.format(f"A{row_index}:P{row_index}", {
            'textFormat': {'fontSize': 10},
            'verticalAlignment': 'MIDDLE'
        })
        worksheet.format(f"A{row_index}", {'horizontalAlignment': 'CENTER'})
        worksheet.format(f"C{row_index}", {'horizontalAlignment': 'CENTER'})
        
        worksheet.format(f"D{row_index}:E{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '#,##0'}
        })
        
        worksheet.format(f"F{row_index}:G{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}
        })
        
        worksheet.format(f"H{row_index}", {
            'horizontalAlignment': 'CENTER',
            'textFormat': {'bold': True}
        })
        
        cell_range = f"H{row_index}"
        if rating == "MUA":
            worksheet.format(cell_range, {
                'textFormat': {'bold': True, 'foregroundColor': {'red': 0.1, 'green': 0.5, 'blue': 0.1}}
            })
        elif rating == "BÁN":
            worksheet.format(cell_range, {
                'textFormat': {'bold': True, 'foregroundColor': {'red': 0.8, 'green': 0.1, 'blue': 0.1}}
            })
        else:
            worksheet.format(cell_range, {
                'textFormat': {'bold': True, 'foregroundColor': {'red': 0.5, 'green': 0.5, 'blue': 0.5}}
            })
            
        worksheet.format(f"I{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '0.0'}
        })
        
        worksheet.format(f"J{row_index}:K{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '0.00'}
        })
        
        worksheet.format(f"L{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}
        })
        
        worksheet.format(f"M{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '#,##0'}
        })
        worksheet.format(f"N{row_index}", {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}
        })
        
        return {"status": "success", "row_updated": row_index}
        
    except Exception as e:
        logger.error(f"Failed to update single ticker to Google Sheets: {e}")
        return {"status": "error", "error": str(e)}


# Tên nhóm ngành (group Excel) → tên đầy đủ cho cột hiển thị
_METHOD_LABEL = {
    "RI_PB": "Thu nhập thặng dư + P/B",
    "DCF": "DCF/FCFF",
    "DCF_EVEBITDA": "DCF + EV/EBITDA",
    "EV_EBITDA": "EV/EBITDA (chuẩn hóa)",
    "PE": "P/E (EPS chuẩn hóa)",
    "PB": "Justified P/B",
    "RNAV": "RNAV (proxy)",
    "SOTP": "SOTP (proxy)",
}


def _confidence_label(result: dict, flags: list[str]) -> str:
    flag_set = set(flags)
    if result.get("fair_value") is None or "NOT_RATEABLE" in flag_set:
        return "Không định giá"
    if result.get("status") == "PARTIAL":
        return "Proxy - không khuyến nghị"
    if any(flag.startswith("STALE_") or flag.startswith("MISSING_") for flag in flags):
        return "Cần cập nhật dữ liệu"
    if {
        "UPSIDE_EXTREME_REVIEW",
        "DOWNSIDE_EXTREME_REVIEW",
        "PROXY_IMPLAUSIBLE",
    }.intersection(flag_set):
        return "Cần kiểm tra"
    if result.get("verified"):
        return "Đã kiểm chứng"
    return "Mô hình ngành - chưa golden test"


def build_vn100_dataframe(results: list) -> pd.DataFrame:
    """Dựng DataFrame bảng định giá VN100 từ kết quả batch.value_all."""
    rows = []
    for r in results:
        if r.get("error"):
            rows.append({
                "Mã": r["ticker"], "Ngành": r.get("group", ""),
                "Phương pháp": _METHOD_LABEL.get(r.get("method", ""), r.get("method", "")),
                "Giá": r.get("price"), "FV": None, "Upside %": None,
                "Khuyến nghị": "LỖI", "Độ tin cậy": "Lỗi", "Cờ": r["error"],
            })
            continue
        flags = r.get("flags", [])
        confidence = _confidence_label(r, flags)
        rows.append({
            "Mã": r["ticker"], "Ngành": r.get("group", ""),
            "Phương pháp": _METHOD_LABEL.get(r["method"], r["method"]),
            "Giá": round(r["price"]) if r.get("price") else None,
            "FV": round(r["fair_value"]) if r.get("fair_value") is not None else None,
            "Upside %": round(r["upside"] * 100, 1) if r.get("upside") is not None else None,
            "Khuyến nghị": r.get("recommendation", "THEO DÕI"),
            "Độ tin cậy": confidence,
            "Cờ": ", ".join(flags),
        })
    df = pd.DataFrame(rows)
    if "Upside %" in df.columns:
        df = df.sort_values("Upside %", ascending=False, na_position="last").reset_index(drop=True)
    return df


def export_vn100_valuations_to_xlsx(
    results: list,
    db: Session,
    output_path: Path,
) -> Path:
    """Xuất workbook local có bảng kết quả và dữ liệu truy vết tối thiểu."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from valuation.ingest.universe import load_vn100_snapshot

    output_path = Path(output_path)
    valuation_df = build_vn100_dataframe(results)
    symbols = [str(result["ticker"]) for result in results]
    quality_rows = []
    for symbol in symbols:
        price_date = (
            db.query(func.max(PricesDaily.trade_date))
            .filter(PricesDaily.ticker == symbol)
            .scalar()
        )
        latest_fin = (
            db.query(
                FinancialsQuarterly.fiscal_year,
                FinancialsQuarterly.fiscal_quarter,
            )
            .filter(
                FinancialsQuarterly.ticker == symbol,
                FinancialsQuarterly.fiscal_quarter > 0,
            )
            .distinct()
            .order_by(
                FinancialsQuarterly.fiscal_year.desc(),
                FinancialsQuarterly.fiscal_quarter.desc(),
            )
            .first()
        )
        sources = [
            row[0]
            for row in db.query(FinancialsQuarterly.source)
            .filter(FinancialsQuarterly.ticker == symbol)
            .distinct()
            .all()
            if row[0]
        ]
        published_rows = (
            db.query(func.count())
            .select_from(FinancialsQuarterly)
            .filter(
                FinancialsQuarterly.ticker == symbol,
                FinancialsQuarterly.published_at.isnot(None),
            )
            .scalar()
        )
        quality_rows.append(
            {
                "Mã": symbol,
                "Ngày giá mới nhất": price_date,
                "Kỳ BCTC mới nhất": (
                    f"{latest_fin[0]}Q{latest_fin[1]}" if latest_fin else None
                ),
                "Nguồn BCTC": ", ".join(sorted(sources)),
                "Số dòng có published_at": int(published_rows or 0),
            }
        )
    quality_df = pd.DataFrame(quality_rows)

    flag_counts = Counter()
    for result in results:
        for flag in result.get("flags", []) or []:
            flag_counts[str(flag).split(":", 1)[0]] += 1
    flag_df = pd.DataFrame(
        [{"Cờ": flag, "Số mã": count} for flag, count in flag_counts.most_common()]
    )

    snapshot = load_vn100_snapshot()
    macro = (
        db.query(MacroSeries)
        .filter(MacroSeries.indicator_code == "TPCP_10Y")
        .order_by(MacroSeries.date.desc())
        .first()
    )
    source_df = pd.DataFrame(
        [
            {
                "Dữ liệu": "Danh mục VN100",
                "Kỳ/Ngày": snapshot["as_of"],
                "Giá trị": "100 mã",
                "Nguồn": snapshot["source"],
                "URL": snapshot["official_reference"],
            },
            {
                "Dữ liệu": "TPCP Việt Nam 10 năm",
                "Kỳ/Ngày": macro.date if macro else None,
                "Giá trị": float(macro.value) if macro else None,
                "Nguồn": macro.source if macro else None,
                "URL": "https://www.hnx.vn/vi-vn/chi-tiet-tin-60023234-0.html",
            },
            {
                "Dữ liệu": "BCTC và giá",
                "Kỳ/Ngày": datetime.date.today(),
                "Giá trị": "Consolidated, VND",
                "Nguồn": "vnstock/VCI",
                "URL": "https://vnstocks.com/",
            },
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        valuation_df.to_excel(writer, sheet_name="VN100_Valuations", index=False)
        quality_df.to_excel(writer, sheet_name="Data_Quality", index=False)
        flag_df.to_excel(writer, sheet_name="Flag_Summary", index=False)
        source_df.to_excel(writer, sheet_name="Sources_Assumptions", index=False)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        bad_fill = PatternFill("solid", fgColor="F4CCCC")
        proxy_fill = PatternFill("solid", fgColor="D9EAD3")
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in worksheet.columns:
                width = min(
                    max(len(str(cell.value or "")) for cell in column) + 2,
                    48,
                )
                worksheet.column_dimensions[column[0].column_letter].width = max(width, 10)

        valuation_ws = writer.book["VN100_Valuations"]
        headers = {cell.value: cell.column for cell in valuation_ws[1]}
        for row in range(2, valuation_ws.max_row + 1):
            confidence = valuation_ws.cell(row, headers["Độ tin cậy"]).value
            target = valuation_ws.cell(row, headers["Độ tin cậy"])
            if confidence == "Không định giá":
                target.fill = bad_fill
            elif confidence == "Proxy - không khuyến nghị":
                target.fill = proxy_fill
            elif confidence in {"Cần cập nhật dữ liệu", "Cần kiểm tra"}:
                target.fill = warning_fill
        for name in ("Giá", "FV"):
            for cell in valuation_ws.iter_cols(
                min_col=headers[name], max_col=headers[name], min_row=2
            ):
                for item in cell:
                    item.number_format = "#,##0"
        for cell in valuation_ws["F"][1:]:
            cell.number_format = '0.0"%"'

    return output_path


def export_vn100_valuations_to_gsheets(results: list, sheet_name: str = "VN100_Valuations"):
    """Xuất bảng định giá VN100 (toàn bộ phương pháp) ra Google Sheets. Gated creds."""
    df = build_vn100_dataframe(results)
    if not settings.google_service_account_json or not settings.google_sheet_master_id:
        logger.warning("Thiếu Google creds — bỏ qua export Sheets, chỉ trả DataFrame.")
        return {"status": "skipped_no_creds", "rows": len(df), "dataframe": df}
    try:
        gc = gspread.service_account(filename=settings.google_service_account_json)
        sh = gc.open_by_key(settings.google_sheet_master_id)
        try:
            ws = sh.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=sheet_name, rows="120", cols="10")
        ws.clear()
        set_with_dataframe(ws, df)
        
        # --- Format đẹp và chuyên nghiệp ---
        num_rows = len(df) + 1
        
        # Format Header
        ws.format('A1:G1', {
            'backgroundColor': {'red': 0.12, 'green': 0.23, 'blue': 0.35},
            'horizontalAlignment': 'CENTER',
            'verticalAlignment': 'MIDDLE',
            'textFormat': {'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0}, 'fontSize': 11, 'bold': True}
        })
        
        # Định dạng toàn bộ bảng
        ws.format(f'A2:G{num_rows}', {
            'textFormat': {'fontSize': 10},
            'verticalAlignment': 'MIDDLE'
        })
        
        # Định dạng cột Giá và FV (số)
        ws.format(f'D2:E{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '#,##0'}
        })
        
        # Định dạng Upside
        ws.format(f'F2:F{num_rows}', {
            'horizontalAlignment': 'RIGHT',
            'numberFormat': {'type': 'NUMBER', 'pattern': '0.0"%"'}
        })
        
        return {"status": "success", "rows": len(df)}
    except Exception as e:
        logger.error(f"Export VN100 thất bại: {e}")
        return {"status": "error", "error": str(e)}
